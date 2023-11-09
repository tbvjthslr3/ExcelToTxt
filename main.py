import openpyxl

# 엑셀 파일 경로 설정
excel_file_path = 'Arbitrage Bot.xlsx'

# 엑셀 파일 열기
workbook = openpyxl.load_workbook(excel_file_path)

# 모든 시트의 데이터를 텍스트 파일로 옮기기
for sheet_name in workbook.sheetnames:
    sheet = workbook[sheet_name]
    txt_file_path = f'{sheet_name}.txt'

    with open(txt_file_path, 'w') as txt_file:
        for row in sheet.iter_rows(values_only=True):
            row_data = ','.join(map(str, row))
            txt_file.write(row_data)
            txt_file.write('\n')

    print(f"'{sheet_name}' 시트의 데이터를 {txt_file_path}로 변환하였습니다.")

print(f'총 {len(workbook.sheetnames)}개의 시트를 처리하였습니다.')
